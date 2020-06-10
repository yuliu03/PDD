import copy
import datetime

from openpyxl import load_workbook, Workbook


#判断是否在活动区间
def checkTimeRange(tmpPddPayTime, tmpBeginTime, tmpEndTime, range):
    try:
        tmpEndTimePlus = tmpEndTime+datetime.timedelta(seconds=range)
    except:
        print("error time")

    result = tmpPddPayTime >= tmpBeginTime and tmpPddPayTime <= tmpEndTimePlus

    return result

#生成excel
#dict[(sku,活动编号，成本价格)]{(商品名称,数量)}
def createExcel(outFilepath,dictInfo):
    newWb = Workbook()

    newSheet = newWb.active

    # 填写字段抬头
    beginRow = 1
    newSheet.cell(row=beginRow, column=1, value="sku")
    newSheet.cell(row=beginRow, column=2, value="活动编号")
    newSheet.cell(row=beginRow, column=3, value="价格")
    newSheet.cell(row=beginRow, column=4, value="商品名称")
    newSheet.cell(row=beginRow, column=5, value="数量")
    beginRow = beginRow + 1

    for key in dictInfo:
        sku = key[0]
        actyCode = key[1]
        basicPrice = key[2]
        proName = dictInfo[key][0]
        num = dictInfo[key][1]

        newSheet.cell(row=beginRow, column=1, value=sku)
        newSheet.cell(row=beginRow, column=2, value=actyCode)
        newSheet.cell(row=beginRow, column=3, value=basicPrice)
        newSheet.cell(row=beginRow, column=4, value=proName)
        newSheet.cell(row=beginRow, column=5, value=num)

        beginRow = beginRow + 1

    newWb.save(outFilepath)

#读取截单数据，输出list[原始单号]
def getJieDanExcel(excelPath,originalOrderNumPos,filterInfo,beginRow):
    workBook = load_workbook(excelPath)
    sheet = workBook.active
    rows = sheet.max_row
    listToReturn = list()

    #遍历每行读取数据
    i = beginRow
    while i <= rows:
        listToReturn.append(str(sheet.cell(row=i, column=originalOrderNumPos).value).strip().strip(filterInfo))
        i = i + 1
    return listToReturn

#读取拼多多表，截单list[原始单号]，根据原始单号过滤，输出list[tuple（订单号，pddIdPos,sku，支付时间，商品总价,商品名称）]
def getPddExcelByFilter(excelPath,beginRow,originalOrderNumList,pddOrderNumPos,pddIdPos,pddSkuPos,pddPayTimePos,productPricePos,proNamePos,filterInfo):
    print("读取截单表中...")
    workBook = load_workbook(excelPath)
    sheet = workBook.active
    rows = sheet.max_row
    listToReturn = list()

    #遍历每一行
    # 遍历每行读取数据
    i = beginRow
    while i <= rows and len(originalOrderNumList)>0:
        print("拼多多数据处理中： "+str(i)+"/"+str(rows))
        #遍历截单的订单号
        j = 0
        while j < len(originalOrderNumList):
            tmpOrderNum = str(sheet.cell(row=i, column=pddOrderNumPos).value.strip().strip(filterInfo))
            # 判断订单号是否在list中
            if str(originalOrderNumList[j]) == tmpOrderNum:
                print("截单数据整理中:"+str(j)+"/"+str(len(originalOrderNumList)))
                tmpPddId = str(sheet.cell(row=i, column=pddIdPos).value).strip().strip(filterInfo)
                tmpSku = str(sheet.cell(row=i, column=pddSkuPos).value).strip().strip(filterInfo)
                tmpPayTime = sheet.cell(row=i, column=pddPayTimePos).value
                tmpProductPrice = float(str(sheet.cell(row=i, column=productPricePos).value).strip().strip(filterInfo))
                tmpProName = str(sheet.cell(row=i, column=proNamePos).value).strip().strip(filterInfo)
                listToReturn.append((tmpOrderNum,tmpPddId,tmpSku,tmpPayTime,tmpProductPrice,tmpProName))

                #从list中删除以处理的数据，缩短
                originalOrderNumList.pop(j)
                #跳出循环
                break
            j = j + 1
        i = i + 1

    print("未处理数据数量：")
    print(originalOrderNumList)
    return listToReturn

#读取活动备案表，输出list[tuple(活动编号,开始时间，结束时间,pddId,sku,活动价，采购价,平台补贴，店铺补贴)]
def getSecctionExcel(excelPath,sheetName,beginRow,secctionCodePos,secctionBeginTimePos,secctionEndTimePos,secctionPddIdPos,secctionSkuPos,secctionPricePos,secctionPurchasePricePos,filterInfo,platformRewardPos,storeRewardPos):
    print("读取活动备案表中...")
    workBook = load_workbook(excelPath)
    # sheet = workBook[sheetName]
    sheet = workBook.active
    rows = sheet.max_row
    columns = sheet.max_column
    listToReturn = list()
    print(rows)


    # 遍历每行读取数据
    i = beginRow
    while i <= rows:
        j = 1
        print("活动备案数据处理中： " + str(i) + "/" + str(rows))
        # while j <= columns:
        #如果是第一次遇到当前活动
        tmpSecctionCode = str(sheet.cell(row=i, column=secctionCodePos).value).strip().strip(filterInfo)
        tmpSecctionBeginTime = sheet.cell(row=i, column=secctionBeginTimePos).value

        if tmpSecctionBeginTime is None:
            tmpSecctionBeginTime =  datetime.datetime(2099, 9, 1)

        tmpSecctionEndTime = sheet.cell(row=i, column=secctionEndTimePos).value
        if tmpSecctionEndTime is None:
            tmpSecctionEndTime =  datetime.datetime(2099, 9, 1)

        listToReturn.append((
        tmpSecctionCode,
        tmpSecctionBeginTime,
        tmpSecctionEndTime,
        str(sheet.cell(row=i, column=secctionPddIdPos).value).strip().strip(filterInfo),
        str(sheet.cell(row=i, column=secctionSkuPos).value).strip().strip(filterInfo),
        sheet.cell(row=i, column=secctionPricePos).value,
        sheet.cell(row=i, column=secctionPurchasePricePos).value,
        sheet.cell(row=i, column=platformRewardPos).value,
        sheet.cell(row=i, column=storeRewardPos).value,
        ))


        i = i + 1

    return listToReturn

#param：活动备案信息， pdd 订单信息，返回list[(订单编号，活动编号，sku，成本价格，商品名称，平台补贴，店铺补贴)]
def runFinalWork(actSecctionList,pdd,range):
    pddSize = len(pdd) #需要处理订单量
    toReturn = list()

    count = 0
    while count < pddSize:
        toNext = 0 #如果已经找到对应商品，就标志

        print("匹配拼多多信息："+str(count)+"/"+str(pddSize))
        ordrNum = pdd[count][0] #订单编号
        tmpPddId = pdd[count][1] #商品id编码
        tmpPddSku  = pdd[count][2] #商品sku编码
        tmpPddPayTime = pdd[count][3] #支付时间
        tmpPddProductAllPrice = pdd[count][4] #商品总价
        tmpPddProName = pdd[count][5] #商品名称


        #遍历所有活动内容
        #list[tuple(活动编号,开始时间，结束时间,pddId,sku,活动价，采购价，平台补贴，店铺补贴)]
        actSheetPos = 1
        for item in actSecctionList:
             actSheetPos = actSheetPos + 1
             if toNext == 1:
                 break

             #活动编码
             tmpActCode = item[0]
             #开始时间
             tmpBeginTime = item[1]
             #结束时间
             tmpEndTime = item[2]
             #pdd 的商品 id
             tmpPddId_actSheet = item[3]
             #商品的sku编码
             tmpSku_actSheet = item[4]
             #活动价
             tmpActPrice = item[5]
             #成本价
             tmpPurchasePrice = item[6]
             # 平台补贴
             tmpPlatformReword = item[7]
             # 店铺补贴
             tmpStoreReword = item[8]

            #判断是否在活动区间
             if checkTimeRange(tmpPddPayTime,tmpBeginTime,tmpEndTime,range):
                #如果是AirPods
                if "AirPods" in tmpPddProName:
                    # 是否是: 同样的sku && 活动价等于商品总价 && 商品id一致
                    if tmpPddId_actSheet == tmpPddId and (
                            tmpSku_actSheet == tmpPddSku and tmpActPrice == tmpPddProductAllPrice):
                        # debug
                        # if  tmpPddSku == "190199227750" and tmpSku_actSheet == "190199227750":
                        #     print()
                        # sku
                        skuToReturn = tmpSku_actSheet
                        # 成本价格
                        basePriceToReturn = tmpPurchasePrice
                        # 商品名称
                        productNameToReturn = tmpPddProName
                        toReturn.append((ordrNum, tmpActCode, skuToReturn, basePriceToReturn, productNameToReturn,
                                         tmpPlatformReword, tmpStoreReword))
                        toNext = 1
                        break
                #如果是iPad
                elif "iPad" in tmpPddProName:
                    # 是否是: 同样的sku
                    if tmpSku_actSheet == tmpPddSku:
                        # 活动价等于商品总价
                        if tmpActPrice == tmpPddProductAllPrice:
                            # sku
                            skuToReturn = tmpSku_actSheet
                            # 成本价格
                            basePriceToReturn = tmpPurchasePrice
                            # 商品名称
                            productNameToReturn = tmpPddProName
                            toReturn.append((ordrNum,tmpActCode, skuToReturn, basePriceToReturn, productNameToReturn,tmpPlatformReword,tmpStoreReword))
                            #跳出活动表循环，准备匹配下一个订单
                            toNext = 1
                            break

                elif "iPhone" in tmpPddProName:
                    # 是否是: 同样的sku && 活动价等于商品总价 && 商品id一致
                    if tmpPddId_actSheet == tmpPddId and (tmpSku_actSheet == tmpPddSku and tmpActPrice == tmpPddProductAllPrice):
                        # debug
                        # if  tmpPddSku == "190199227750" and tmpSku_actSheet == "190199227750":
                        #     print()
                        # sku
                        skuToReturn = tmpSku_actSheet
                        # 成本价格
                        basePriceToReturn = tmpPurchasePrice
                        # 商品名称
                        productNameToReturn = tmpPddProName
                        toReturn.append((ordrNum,tmpActCode, skuToReturn, basePriceToReturn, productNameToReturn,tmpPlatformReword,tmpStoreReword))
                        toNext = 1
                        break
                #如果是其他
                else:
                    print("其他商品："+tmpPddProName)

        count = count + 1
    return toReturn

#商品分类汇总计算数量。 param: list[(订单编号，活动编号，sku，成本价格，商品名称)]; return: dict[(sku,活动编号，成本价格)]{(商品名称,数量)}
def clssify(purchaseList):
    dictToReturn = dict()
    for item in purchaseList:
        tmpActCode = item[1]
        tmpSku = item[2]
        tmpBasicPrice = item[3]
        tmpProName = item[4]

        if (tmpSku,tmpActCode,tmpBasicPrice) not in dictToReturn:
            dictToReturn[(tmpSku,tmpActCode,tmpBasicPrice)] = (tmpProName,1)
        else:
            oldInfo = dictToReturn[(tmpSku,tmpActCode,tmpBasicPrice)]
            dictToReturn[(tmpSku, tmpActCode, tmpBasicPrice)] = (oldInfo[0], oldInfo[1]+1)

    return dictToReturn

#根据订单号计算不符合标准订单,param  : list[tuple（订单号，pddIdPos,sku，支付时间，商品总价,商品名称）
def checkList(pddListBefore,pddListAfter):
    listToReturn = list()
    for oldItem in pddListBefore:
        #判断老的信息是否在新的列表里出现
        isTreated = 0
        pddListAfterSize = len(pddListAfter)
        index = 0
        while index < pddListAfterSize:
            newItem = pddListAfter[index]
            if newItem[0] == oldItem[0]:
                isTreated = 1
                break
            index = index + 1
        if isTreated == 0:
            listToReturn.append(oldItem)

    return listToReturn

#标记pdd Excel 中订单对应的活动编码:param: 采购单，pdd文件地址，pdd订单号位置，输出位置，pdd商家实际收入位置
def markActCode(purchaseExcel, pddExcelPath,ordNumPos,outputPath,pddRealIncomePos):
    print("读取拼多多明细表中...")
    workBook = load_workbook(pddExcelPath)
    sheet = workBook.active
    rows = sheet.max_row
    cols = sheet.max_column
    toWriteActPos  = cols + 1
    toWritePlatformRewordPos = cols + 2
    toWriteStoreRewordPos = cols + 3
    toWriteBasicPrice = cols + 4
    toWriteFinalReword =cols + 5
    toWriteRealIncome = cols + 6
    #拼多多对账明细的店铺补贴字段位置
    pdd_storeRewordPos = 5

    sheet.cell(row=1, column=toWriteActPos, value="活动备案表_活动编码")
    sheet.cell(row=1, column=toWritePlatformRewordPos, value="活动备案表_平台补贴")
    sheet.cell(row=1, column=toWriteStoreRewordPos, value="活动备案表_店铺补贴")
    sheet.cell(row=1, column=toWriteBasicPrice, value="活动备案表_成本")
    sheet.cell(row=1, column=toWriteFinalReword, value="补贴金额")
    sheet.cell(row=1, column=toWriteRealIncome, value="千6费用")
    beginRow = 2

    i = beginRow
    while i <= rows:
        print("填充补贴、活动编号等数据到拼多多明细表："+str(i)+"/"+str(rows))
        ordNum = str(sheet.cell(row=i,column=ordNumPos).value.strip().strip(filterInfo))
        #list[(订单编号，活动编号，sku，成本价格，商品名称，平台补贴，店铺补贴)]
        for item in purchaseExcel:
            # sheet.cell(row=i, column=toWritePos, value="item[1]")
            if item[0] == ordNum:
                sheet.cell(row=i, column=toWriteActPos,value=item[1])
                sheet.cell(row=i, column=toWriteBasicPrice, value=item[3])
                sheet.cell(row=i, column=toWritePlatformRewordPos, value=item[5])
                sheet.cell(row=i, column=toWriteStoreRewordPos, value=item[6])
                realIncome = float(sheet.cell(row=i, column=pddRealIncomePos).value)
                sheet.cell(row=i, column=toWriteRealIncome, value=(realIncome * 0.006))
                #活动编号不为空
                if item[1] is not None and item[1] != "":
                    #店铺补贴不为空 && 店铺补贴等于多多对账明细的店铺补贴
                    if item[6] is not None and str(item[6]) == str(sheet.cell(row=i, column=pdd_storeRewordPos).value):
                        sheet.cell(row=i, column=toWriteFinalReword, value=float(item[5]) + float(item[6])) #店铺补贴 + 平台补贴
                    #非空校验
                    elif sheet.cell(row=i, column=pdd_storeRewordPos).value is not None and \
                            sheet.cell(row=i, column=pdd_storeRewordPos).value != "":
                            #判断金额
                            if float(str(sheet.cell(row=i, column=pdd_storeRewordPos).value)) == 0:
                                sheet.cell(row=i, column=toWriteFinalReword, value=float(item[6])) #平台补贴
                break
        i = i + 1
    workBook.save(outputPath)

#生成未符合标准订单表
def creatFaultListExcel(faultList,faultOrderExcelPath):
    if len(faultList)==0:
        return 0
    print("开始生成未符合标准订单表")
    newWb = Workbook()

    newSheet = newWb.active

    # 填写字段抬头
    beginRow = 1
    newSheet.cell(row=beginRow, column=1, value="订单号")
    newSheet.cell(row=beginRow, column=2, value="pdd Id")
    newSheet.cell(row=beginRow, column=3, value="商品sku编码")
    newSheet.cell(row=beginRow, column=4, value="支付时间")
    newSheet.cell(row=beginRow, column=5, value="支付金额")
    newSheet.cell(row=beginRow, column=6, value="商品名称")

    beginRow = beginRow + 1
    indexPos = 0
    size = len(faultList)

    while indexPos < size:
        item = faultList[indexPos]
        newSheet.cell(row=beginRow, column=1, value=item[0])
        newSheet.cell(row=beginRow, column=2, value=item[1])
        newSheet.cell(row=beginRow, column=3, value=item[2])
        newSheet.cell(row=beginRow, column=4, value=item[3])
        newSheet.cell(row=beginRow, column=5, value=item[4])
        newSheet.cell(row=beginRow, column=6, value=item[5])
        beginRow = beginRow + 1
        indexPos = indexPos + 1

    newWb.save(faultOrderExcelPath)

#处理未符合标准订单，param： list[tuple（订单号，pddIdPos,sku，支付时间，商品总价,商品名称）; toReturn 是最终返回的采购订单结果
def dealFaultData(originalFaultList,actData):
    toReturn = list()
    tmpFaultList = copy.copy(originalFaultList)
    size = len(tmpFaultList)
    index = 0
    #遍历所有的未符合标准订单
    while index < size:
    # for item in faultList:
        item = originalFaultList[index]
        #初始化时间差
        lessDistance = 99999999
        #临时结果
        tmpInf = None
        #临时数据指针
        tmpIndex = None
        #初始化订单号
        ordrNum = item[0]
        #初始化pdd id
        tmpPddId = item[1]
        #初始化sku
        tmpPddSku = item[2]
        # 获取支付时间
        tmpPayTime = item[3]
        #初始化订单总价
        tmpPddProductAllPrice = item[4]
        #初始化商品名称
        tmpPddProName = item[5]

        #遍历活动信息
        for eachAct in actData:
            #初始化活动表信息
            # 活动编码
            tmpActCode = eachAct[0]
            # 开始时间
            tmpBeginTime = eachAct[1]
            # 结束时间
            tmpEndTime = eachAct[2]
            # pdd 的商品 id
            tmpPddId_actSheet = eachAct[3]
            # 商品的sku编码
            tmpSku_actSheet = eachAct[4]
            # 活动价
            tmpActPrice = eachAct[5]
            # 成本价
            tmpPurchasePrice = eachAct[6]
            # 平台补贴
            tmpPlatformReword = eachAct[7]
            # 店铺补贴
            tmpStoreReword = eachAct[8]

            #debugger


            #判断基本条件（）
            #支付时晚于活动开始时间
            if tmpPayTime > tmpBeginTime:
                # 如果是AirPods
                if "AirPods" in tmpPddProName:
                    # 是否是: 同样的sku && 活动价等于商品总价 && 商品id一致
                    if tmpPddId_actSheet == tmpPddId and (
                            tmpSku_actSheet == tmpPddSku and tmpActPrice == tmpPddProductAllPrice):

                        skuToReturn = tmpSku_actSheet
                        # 成本价格
                        basePriceToReturn = tmpPurchasePrice
                        # 商品名称
                        productNameToReturn = tmpPddProName

                        #距离活动结束时间时长
                        tmpDistance = tmpPayTime.timestamp() - tmpEndTime.timestamp()

                        if lessDistance > tmpDistance:
                            tmpInf = (ordrNum, tmpActCode, skuToReturn, basePriceToReturn, productNameToReturn,
                                  tmpPlatformReword, tmpStoreReword)
                            lessDistance = tmpDistance
                            tmpIndex = index

                # 如果是iPad
                elif "iPad" in tmpPddProName:
                    # 是否是: 同样的sku
                    if tmpSku_actSheet == tmpPddSku:
                        # 活动价等于商品总价
                        if tmpActPrice == tmpPddProductAllPrice:
                            # sku
                            skuToReturn = tmpSku_actSheet
                            # 成本价格
                            basePriceToReturn = tmpPurchasePrice
                            # 商品名称
                            productNameToReturn = tmpPddProName

                            # 距离活动结束时间时长
                            tmpDistance = tmpPayTime.timestamp() - tmpEndTime.timestamp()

                            if lessDistance > tmpDistance:
                                tmpInf = (ordrNum, tmpActCode, skuToReturn, basePriceToReturn, productNameToReturn,
                                          tmpPlatformReword, tmpStoreReword)
                                lessDistance = tmpDistance
                                tmpIndex = index

                elif "iPhone" in tmpPddProName:
                    # 是否是: 同样的sku && 活动价等于商品总价 && 商品id一致
                    if tmpPddId_actSheet == tmpPddId and (
                            tmpSku_actSheet == tmpPddSku and tmpActPrice == tmpPddProductAllPrice):

                        skuToReturn = tmpSku_actSheet
                        # 成本价格
                        basePriceToReturn = tmpPurchasePrice
                        # 商品名称
                        productNameToReturn = tmpPddProName

                        # 距离活动结束时间时长
                        tmpDistance = tmpPayTime.timestamp() - tmpEndTime.timestamp()

                        if lessDistance > tmpDistance:
                            tmpInf = (ordrNum, tmpActCode, skuToReturn, basePriceToReturn, productNameToReturn,
                                      tmpPlatformReword, tmpStoreReword)
                            lessDistance = tmpDistance
                            tmpIndex = index

                # 如果是其他
                else:
                    print("其他商品：" + tmpPddProName)

        if tmpInf is not None:
            toReturn.append(tmpInf)
            tmpFaultList.remove(item)
            print(tmpFaultList)

        index = index + 1
    return toReturn,tmpFaultList

jieDanExcelPath = "C:/Users/admin/Desktop/截单/jiedan.xlsx"
jieDanOriginalOrderNumPos = 3
beginRow = 2
filterInfo = "\t"

jieDanList = getJieDanExcel(jieDanExcelPath,jieDanOriginalOrderNumPos,filterInfo,beginRow)
jieDanListSize = len(jieDanList)

pddExcelPath = "C:/Users/admin/Desktop/截单/pddTest.xlsx"
originalOrderNumList = jieDanList
pddOrderNumPos = 2
pddSkuPos = 33
pddPayTimePos = 23
productPricePos = 4
pddIdPos = 29
proNamePos = 1
pddFilterInfo = "\t"
beginRow = 2
pddExcel = getPddExcelByFilter(pddExcelPath,beginRow,originalOrderNumList,pddOrderNumPos,pddIdPos,pddSkuPos,pddPayTimePos,productPricePos,proNamePos,pddFilterInfo)


pddFinanceExcelPath = "C:/Users/admin/Desktop/截单/huodong.xlsx"
sheetName = "活动备案"
secctionCodePos = 3
secctionBeginTimePos = 4
secctionEndTimePos = 5
secctionSkuPos = 6
secctionPricePos = 11
secctionPurchasePricePos = 9
secctionPddIdPos = 2
platformRewardPos = 13
storeRewardPos = 14
filterInfo = "\t"
beginRow = 2
financeDict = getSecctionExcel(pddFinanceExcelPath,sheetName,beginRow,secctionCodePos,secctionBeginTimePos,secctionEndTimePos,secctionPddIdPos,secctionSkuPos,secctionPricePos,secctionPurchasePricePos,filterInfo,platformRewardPos,storeRewardPos)
print(financeDict)


#活动浮动区间 秒
range = 0
purchaseExcel=runFinalWork(financeDict,pddExcel,range)

#检查未符合标准的订单
faultList = checkList(pddExcel,purchaseExcel)

#处理未在活动时间内支付的订单
faultListDealed,filterFaultList = dealFaultData(faultList,financeDict)

#把未在活动时间内支付的订单的处理结果添加进正常订单结果
purchaseExcel.extend(faultListDealed)

#分类汇总数据
finalInfo = clssify(purchaseExcel)

#生成未符合标准商品（疑似未在活动内的订单）
faultOrderExcelPath = "C:/Users/admin/Desktop/截单/未符合标准商品（疑似未在活动内的订单）.xlsx"
creatFaultListExcel(filterFaultList,faultOrderExcelPath)

outputPath = "C:/Users/admin/Desktop/截单/pddTestWithActCode.xlsx"
pddRealIncomePos = 12
markActCode(purchaseExcel,pddExcelPath,pddOrderNumPos,outputPath,pddRealIncomePos)

print("________备案表获取结果 financeDict________________")
print(len(financeDict))
print(financeDict)
print()

print("_________截单表内不符合拼多多订单表内信息 jieDanList______________")
print(jieDanList)
print()

print("_________拼多多过滤截单后数据 pddExcel_____________")
print(len(pddExcel))
print(pddExcel)
print()

print("____________对比备案表结果 finalExcel________")
print(len(purchaseExcel))
print(purchaseExcel)
print()

print("______________未符合标准商品（疑似未在活动内的订单）:________")
print(len(filterFaultList))
print(filterFaultList)
print()

print("______________分类汇总最终结果 finalInfo________")
print(len(finalInfo))
print(finalInfo)
print()


outFilepath = "C:/Users/admin/Desktop/截单/outPut.xlsx"
createExcel(outFilepath,finalInfo)