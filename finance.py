from openpyxl import load_workbook, Workbook

#inputMatriz1 输入的数据1，类似vlookup的第一个参数
#inputMatriz2 输入的数据2，类似vlookup的第二个参数
#otherCondColDict1 类似vlookup 的第三个参数，但是对参数inputMatriz1使用
#otherCondColDict2 类似vlookup 的第三个参数

def specialVlookUp(inputMatriz1, condColPos1, otherCondColDict1, inputMatriz2, condColPos2, otherCondColDict2, obtainInfo, specialVLookUpType="normal"):
    toReturn = list()
    if specialVLookUpType == "normal":
        print(specialVLookUpType)
        for i in inputMatriz1[condColPos1]:
            xRow = containValue(i, condColPos2, inputMatriz2)
            # do something with info in xRow
            if(xRow is not None):
                print(xRow)

    return obtainInfo

#单列去重复值结果，返回list
def noRepeat(colPos,matriz):
    toReturn = list()
    for i in matriz[colPos]:
        if str(i) not in toReturn:
            toReturn.append(i)

    return toReturn

#返回一个字典：
def classify(value,colPos,matriz,toReturn):
    if toReturn.has_key(str(value)):
        return -1

    size = len(matriz[colPos])
    listValues = list()
    index = 0
    while index < size:
        if str(value) == matriz[colPos][index]:
            listValues.append(index)
        index = index + 1
    toReturn[str(value)] = listValues

    return toReturn


def containValue(value, colPos, matriz):
    for i in matriz:
        if str(value) == i[colPos]:
            return i
    return None


def getRowPos(value,colPos,sheet):
    rows = sheet.max_row
    i = 1
    while i <= rows:
        if str(sheet.cell(row=i, column=colPos).value) == str(value):
            return i
        i = i + 1


def getDictFromExcel(excelPath,sheetName):
    workBook = load_workbook(excelPath)
    sheet = workBook[sheetName]
    rows = sheet.max_row
    columns = sheet.max_column
    toReturn = list()
    i = 1
    while i <= rows:
        j = 1
        tmpList = list()
        while j <= columns:
            tmpList.append(sheet.cell(row=i, column=j).value)
            j = j + 1
        toReturn.append(tmpList)
        i = i + 1
    return toReturn

#读取文件路径
inputFilepath = "C:/Users/admin/Desktop/test.xlsx"

#文件定位
##活动备案表#########################
actNameList = list()

actName = "活动备案"

#活动类型
an_actTypeName = "活动类型"
an_actType = 1
#活动编号
an_actCode = "活动编号"
an_actPos = 2
#提报时间
an_actStartDateName = "提报时间"
an_actStartDatePos = 3
#结束时间
an_actEndDateName = "结束时间"
an_actEndDatePos = 3
#商品sku
an_skuName = "商品sku"
an_skuPos = 4
#商品名称
an_productName = "商品名称"
an_productNamePos = 5
#商品规格
an_productDetailName = "商品规格"
an_productDetailPos = 6
#采购价
an_purchasePriceName = "采购价"
an_purchasePricePos = 7
#拼多多-结算价
an_pddFinalPriceName = "拼多多-结算价"
an_pddFinalPricePos = 8
#活动价
an_pddActPriceName = "活动价"
an_pddActPricePos = 9
#实收价
an_realReceivePriceName = "实收价"
an_realReceivePricePos = 10
#平台补贴
an_plateformRewordName = "平台补贴"
an_plateformReword = 11
#店铺补贴
an_storeRewordName = "店铺补贴"
an_storeRewordPos = 12
#销售模式
an_saleTypeName = "销售模式"
an_saleTypePos = 13
#活动数量
an_actNumName = "活动数量"
an_actNumPos = 14
#销售数量
an_saleNumName = "销售数量"
an_saleNumPos = 15
#补贴金额
an_rewordPriceName = "补贴金额"
an_rewordPricePos = 16
#可退服务费
an_returnServicePriceName = "可退服务费"
an_returnServicePricePos = 17




##采购明细表######################
purchaseName="拼多多订单明细-对账"
purchaseNameList = list()

#商品
pn_productName = "商品"
pn_productPos = 1
#订单号
pn_orderIdName = "订单号"
pn_orderIdPos = 2
#订单状态
pn_orderStateName = "订单状态"
pn_orderStatePos = 3
#商品总价（元）
pn_productTotalPriceName = "商品总价（元）"
pn_productTotalPriceId = 4
#店铺优惠折扣(元)
pn_storeDiscountPriceName = "店铺优惠折扣(元)"
pn_storeDiscountPricePos = 5
#平台优惠折扣(元)
pn_platformDiscountPriceName = "平台优惠折扣(元)"
pn_platformDiscountPricePos = 6
#邮费(元)
pn_postPriceName = "邮费(元)"
pn_postPricePos = 7
#上门安装费(元)
pn_homeServicePriceName = "上门安装费(元)"
pn_homeServicePricePos = 8
#送货入户费(元)
pn_homeSendPriceName = "送货入户费(元)"
pn_homeSendPricePos = 9
#送货入户并安装费(元)
pn_homeSendExtPriceName = "送货入户并安装费(元)"
pn_homeSendExtPricePos = 10
#用户实付金额(元)
pn_clientRealPayName = "用户实付金额(元)"
pn_clientRealPayId = 11
#商品数量(件)
pn_productNumName = "商品数量(件)"
pn_productNumPos = 12
#身份证姓名
pn_idCardNameName = "身份证姓名"
pn_idCardNamePos = 13
#身份证号码
pn_idCardNum = "身份证号码"
pn_idCardNumPos = 14
#收货人
pn_receivePersonName = "收货人"
pn_receivePersonPos = 15
#手机
pn_telNumName = "手机"
pn_telNumPos = 16
#是否异常
pn_isNormalName = "是否异常"
pn_isNormalPos = 17
#省
pn_provinceName = "省"
pn_provincePos = 18
#市
pn_cityName = "市"
pn_cityPos = 19
#区
pn_districtName = "区"
pn_districtPos = 20
#街道
pn_roadName = "街道"
pn_roadPos = 21
#支付时间
pn_payTimeName = "支付时间"
pn_payTimePos = 22
#拼单成功时间
pn_madeBigOrderOkName = "支付时间"
pn_madeBigOrderOkPos = 23
#订单确认时间
pn_orderOkName = "订单确认时间"
pn_orderOkPos = 24
#承诺发货时间
pn_promiseSendTimeName = "承诺发货时间"
pn_promiseSendTimePos = 25
#发货时间
pn_sendProductTimeName = "发货时间"
pn_sendProductPos = 26
#确认收货时间
pn_confirmReceiveTimeName = "确认收货时间"
pn_confirmReceiveTimePos = 27
#商品id
pn_productNameIdName = "商品id"
pn_productNameIdPos = 28
#商品规格
pn_productDetailName = "商品规格"
pn_productDetailPos = 29
#用户购买手机号
pn_productTelNumName = "用户购买手机号"
pn_productTelNumPos = 30
#样式ID
pn_styleIdName = "样式ID"
pn_styleIdPos = 31
#商家编码-SKU维度
pn_sailerSkuCodeName = "商家编码-SKU维度"
pn_sailerSkuCodePos = 32
#商家编码-商品维度
pn_sailerProductCodeName = "商家编码-商品维度"
pn_sailerProductCodePos = 33
#快递单号
pn_postCodeName = "快递单号"
pn_postCodePos = 34
#快递公司
pn_postCompanyName = "快递公司"
pn_postCompanyPos = 35
#海淘清关订单号
pn_declareOrderIdName = "海淘清关订单号"
pn_declareOrderIdPos = 36
#支付ID
pn_payIdName = "支付ID"
pn_payIdPos = 37
#支付方式
pn_payTypeName = "支付方式"
pn_payTypePos = 38
#是否抽奖或0元试用
pn_isFreePriceTryName = "是否抽奖或0元试用"
pn_isFreePriceTryPos = 39
#是否顺丰加价
pn_isSFExtPriceName = "是否顺丰加价"
pn_isSFExtPricePos = 40
#商家备注
pn_salerRemarkName = "商家备注"
pn_salerRemarkPos = 41
#售后状态
pn_afterSaleServiceName = "售后状态"
pn_afterSaleServicePos = 42
#买家留言
pn_buyerRemarkName = "买家留言"
pn_buyerRemarkPos = 43
#关联货品编码
pn_productRelativeCodeName = "关联货品编码"
pn_productRelativeCodePos = 44
#货品名称
pn_product2name = "货品名称"
pn_product2Pos = 45
#货品类型
pn_product2TypeName = "货品类型"
pn_product2TypePos = 46
#子货品
pn_childProductName = "子货品"
pn_childProductPos = 47
#仓库名称
pn_wmName = "仓库名称"
pn_wmPos = 48
#仓库所在地址
pn_wmAddName = "仓库所在地址"
pn_wmAddPos = 49
#是否门店自提
pn_isClientSelfGetName = "是否门店自提"
pn_isClientSelfGetPos = 59
#门店名称
pn_storeNameName = "门店名称"
pn_storeNamePos = 60
#门店自定义编码
pn_storeCodeName = "门店自定义编码"
pn_storeCodePos = 61
#旅行类信息
pn_traveInfName = "旅行类信息"
pn_traveInfPos = 62
#已发货
pn_hasSentStateName = "已发货"
pn_hasSentStatePos = 63
#采购订单号
pn_purchaseIdName = "采购订单号"
pn_purchaseIdPos = 64
#手续费
pn_commissionName = "手续费"
pn_commissionPos = 65
#0.6 费用
pn_sixExtPriceName = "平台千六费用"
pn_sixExtPricePos = 66
#实收
pn_realGainName = "实收"
pn_realGainPos = 67
#活动标号
pn_actCodeName = "活动标号"
pn_actCodePos = 68
#确认收入
pn_confirmGainName = "确认收入"
pn_confirmGainPos = 69

print(getDictFromExcel(inputFilepath,actName))

print(getDictFromExcel(inputFilepath,purchaseName))







