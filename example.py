from tkinter import Tk, Button, Entry, Label, IntVar, Checkbutton, Radiobutton
from tkinter import filedialog


from openpyxl import load_workbook, Workbook


# row | 1 - n
# col | 1 - n
# 获取某一单元格内容
def readExcel(row, col, sheet):
    return sheet.cell(row=row, column=col).value


# 根据名称获取sheet
def getSheet(wb, sheetName="Sheet0"):
    return wb[sheetName]


# 获取表格的总行数和总列数
def getRowsClosNum(sheet):
    rows = sheet.max_row
    columns = sheet.max_column
    return rows, columns


# 获取excel文件
def getWorkBook(path):
    return load_workbook(path)


# 获取某个单元格的值
def getCellValue(sheet, row, column):
    cellvalue = None
    try:
        cellvalue = sheet.cell(row=row, column=column).value
    except:
        print("位置： "+str(row)+", "+str(column)+" 读取失败")

    return cellvalue


# 获取某列的所有值
def getColValues(sheet, column):
    rows = sheet.max_row
    columndata = []
    for i in range(1, rows + 1):
        cellvalue = sheet.cell(row=i, column=column).value
        columndata.append(cellvalue)
    return columndata


# 获取某行所有值
def getRowValues(sheet, row):
    columns = sheet.max_column
    rowdata = []
    for i in range(1, columns + 1):
        cellvalue = sheet.cell(row=row, column=i).value
        rowdata.append(cellvalue)
    return rowdata


#读取财务文件
def readInputSheet(beginRow, sheet, productName, purchase_price, group_price, non_group_price, stock,sku):
    toReturndict = dict()
    x_row = beginRow
    while x_row < sheet.max_row:
        tmp_sku = str(getCellValue(sheet, x_row, productName)).replace(" ", "")
        if tmp_sku in toReturndict:
            raise Exception(print(str(tmp_sku) + ": 重复出现"))
        else:
            eachValueList = []
            #获取采购价
            eachValueList.append(float(getCellValue(sheet,x_row,purchase_price)))
            #获取拼团价
            eachValueList.append(float(getCellValue(sheet,x_row,group_price)))
            #获取非拼团价格
            eachValueList.append(float(getCellValue(sheet,x_row,non_group_price)))
            #获取库存
            eachValueList.append(float(getCellValue(sheet,x_row,stock)))
            # 获取库存
            eachValueList.append(str(getCellValue(sheet, x_row, sku)))

            toReturndict[tmp_sku] =  eachValueList

        x_row = x_row + 1

    return toReturndict

#对比财务文件,param: {sku:[ purchase_price,group_price,non_group_price,stock]}
def compare(todayDict,yesterdayDict,each_purchase_price,each_group_price):
    quickSellDict = dict()
    nonQuickSellDict = dict()


    for key in todayDict:
        if key in yesterdayDict:
            try:
                resultList = []
                #如果今天拼团价大于昨天的拼团价格
                if todayDict[key][each_group_price] - yesterdayDict[key][each_group_price] > 0:
                    quickSellDict[key] = todayDict[key]
                    #加入指数
                    if todayDict[key][each_purchase_price] == 0:
                        nonQuickSellDict[key].append("0 不能为除数")
                    else:
                        index = ((todayDict[key][each_group_price] - todayDict[key][each_purchase_price]) / todayDict[key][
                            each_purchase_price])
                        quickSellDict[key].append(index)

                else:
                    nonQuickSellDict[key] = todayDict[key]
                    if todayDict[key][each_purchase_price] == 0:
                        nonQuickSellDict[key].append("0 不能为除数")
                    else:
                        index = ((todayDict[key][each_group_price] - todayDict[key][each_purchase_price]) / todayDict[key][
                            each_purchase_price])
                        nonQuickSellDict[key].append(index)
            except:
                raise Exception("product Name："+key+" 报错")
        else:
            raise Exception("product Name："+key+" 未在昨天的单子里出现")

    return quickSellDict , nonQuickSellDict

def writeSheet(newSheet,inputDict,beginRow):
    tmp_row = beginRow
    for key in inputDict.keys():
        i = 0
        newSheet.cell(row=tmp_row, column=i+1, value=key)  # sku
        i = i + 1
        while i <= len(inputDict[key]):
            newSheet.cell(row=tmp_row, column=i+1, value=inputDict[key][i-1])
            i = i + 1
        tmp_row = tmp_row + 1

def writeTitle(newSheet,titleList,beginRow=1):
    i = 0
    while i < len(titleList):
        newSheet.cell(row=beginRow, column=i + 1, value=titleList[i])
        i = i + 1


outFilepath = "C:/Users/admin/Desktop/zangweioutput.xlsx"

#文件列的位置
productName = 1
purchase_price = 3
group_price = 6
non_group_price = 5
stock = 8
sku = 9

#文件内容起始行
todaySheetBeginRow = 2
yesterdaySheetBeginRow = 2

#文件路径
# todayWorkBookPath = "C:/Users/admin/Desktop/zwtest/副本0311-订单宏佰珈报价-拼多多建议价1.xlsx"
# yesterdayWorkBookPath = "C:/Users/admin/Desktop/zwtest/副本0310-3点后订单安彬报价-拼多多建议价.xlsx"

#输出文件的sheet的名称
quickSellSheetName = "急上架"
nonQuickSellSheetName = "非急上架"

#输出文件sheet 的顺序
quickSellIndex = 0
nonQuickSellIndex = 1

#输出结果，虚拟内容的结果值的位置
each_purchase_price = 0
each_group_price = 1
each_non_group_price = 2
each_stock = 3
each_sku = 4

#title内容
quickSellSheetTitle = ["商品", "采购价", "团购价", "非团购价", "数量","sku", "指数"]
nonQuickSellSheetTitle = ["商品", "采购价", "团购价", "非团购价", "数量","sku", "指数"]


def doWork(todayWorkBookPath,yesterdayWorkBookPath):
    todayWorkBook = getWorkBook(todayWorkBookPath)
    todaySheet = getSheet(todayWorkBook, todayWorkBook.sheetnames[0])

    yesterdayWorkBook = getWorkBook(yesterdayWorkBookPath)
    yesterdaySheet = getSheet(yesterdayWorkBook, yesterdayWorkBook.sheetnames[0])

    #获取文件内容
    todayDict = readInputSheet(todaySheetBeginRow, todaySheet,productName, purchase_price,group_price,non_group_price,stock,sku )
    yesterdayDict = readInputSheet(yesterdaySheetBeginRow, yesterdaySheet,productName, purchase_price,group_price,non_group_price,stock,sku )

    #获取对比文件结果
    quickSellDict, nonQuickSellDict=compare(todayDict,yesterdayDict,each_purchase_price,each_group_price)

    #新建excel
    newWb = Workbook()
    quickSellSheet = newWb.create_sheet(title=quickSellSheetName, index=quickSellIndex)
    writeSheet(quickSellSheet, quickSellDict, 2)
    writeTitle(quickSellSheet,quickSellSheetTitle)

    nonQuickSellSheet = newWb.create_sheet(title=nonQuickSellSheetName, index=nonQuickSellIndex)
    writeSheet(nonQuickSellSheet, nonQuickSellDict, 2)
    writeTitle(nonQuickSellSheet,nonQuickSellSheetTitle)


    newWb.remove(newWb["Sheet"])
    newWb.save(outFilepath)


def getInputPath(file_input_path_ui):
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()

    file_input_path_ui.insert(0,eval(repr(file_path.replace('\\\\', '/'))))




if __name__ == "__main__":
    root = Tk()
    Label(root, text="今日文件: ").grid(row=0, column=0)
    todayFile_input_path_ui = Entry(root, width=60)
    todayFile_input_path_ui.grid(row=0, column=1)

    Label(root, text="昨日文件: ").grid(row=1, column=0)
    yesterdayFile_input_path_ui = Entry(root, width=60)
    yesterdayFile_input_path_ui.grid(row=1, column=1)

    todayPathButton = Button(root, text='获取当天文件路径', command=lambda :getInputPath(todayFile_input_path_ui), width=15)
    todayPathButton.grid(row=2, column=0)

    yesterdayPathButton = Button(root, text='获取昨天文件路径', command=lambda :getInputPath(yesterdayFile_input_path_ui), width=15)
    yesterdayPathButton.grid(row=2, column=1)

    createResultButton = Button(root, text='保存', command=lambda: doWork(todayFile_input_path_ui.get(),yesterdayFile_input_path_ui.get()),
                                 width=15)
    createResultButton.grid(row=2, column=2)

    root.mainloop()