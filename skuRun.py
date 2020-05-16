from openpyxl import load_workbook, Workbook

#inputMatriz1 输入的数据1，类似vlookup的第一个参数
#inputMatriz2 输入的数据2，类似vlookup的第二个参数
#otherCondColDict1 类似vlookup 的第三个参数，但是对参数inputMatriz1使用
#otherCondColDict2 类似vlookup 的第三个参数

def specialVlookUp(inputMatriz1, condColPos1, otherCondColDict1, inputMatriz2, condColPos2, otherCondColDict2, obtainInfo, specialVLookUpType="normal"):
    toReturn = list()
    if specialVLookUpType == "normal":
        print(specialVLookUpType)
        for i in inputMatriz1[condColPos1]:
            xRow = containValue(i, condColPos2, inputMatriz2)
            # do something with info in xRow
            if(xRow is not None):
                print(xRow)

    return obtainInfo



#单列去重复值结果，返回list
def noRepeat(colPos,matriz):
    toReturn = list()
    for i in matriz[colPos]:
        if str(i) not in toReturn:
            toReturn.append(i)

    return toReturn

#按照数值在列中汇总，返回一个字典：
def classify(value,colPos,matriz,toReturn):
    if toReturn.has_key(str(value)):
        return -1

    size = len(matriz[colPos])
    listValues = list()
    index = 0
    while index < size:
        if str(value) == matriz[colPos][index]:
            listValues.append(index)
        index = index + 1
    toReturn[str(value)] = listValues

    return toReturn


#在matriz里面判断，回传value的行位置-1
def containValue(value, colPos, matriz):
    for i in matriz:
        if str(value) == i[colPos]:
            return i
    return None

#在sheet里面判断，回传value的行位置
def getRowPos(value,colPos,sheet):
    rows = sheet.max_row
    i = 1
    while i <= rows:
        if str(sheet.cell(row=i, column=colPos).value) == str(value):
            return i
        i = i + 1

#读取sheet内容，回传matriz
def getDictFromExcel(excelPath,sheetName):
    workBook = load_workbook(excelPath)
    sheet = workBook[sheetName]
    rows = sheet.max_row
    columns = sheet.max_column
    toReturn = list()
    i = 1
    while i <= rows:
        j = 1
        tmpList = list()
        while j <= columns:
            tmpList.append(sheet.cell(row=i, column=j).value)
            j = j + 1
        toReturn.append(tmpList)
        i = i + 1
    return toReturn

#读取文件路径
inputFilepath = "C:/Users/admin/Desktop/test.xlsx"